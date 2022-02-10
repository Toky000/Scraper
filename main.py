from bs4 import BeautifulSoup
import requests
import nltk
from nltk.corpus import stopwords
import openpyxl
from nltk import sent_tokenize

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:55.0) Gecko/20100101 Firefox/55.0',
}

input_file = "./resources/Input.xlsx"
output_raw_file = "./resources/Output Data Structure.xlsx"
output_file = "./output/Output Data Structure.xlsx"
stopwords_generic = "./resources/StopWords_Generic.txt"
positive_dictionary = "./resources/PositiveWords.txt"
negative_dictionary = "./resources/NegativeWords.txt"

with open(positive_dictionary, 'r') as posfile:
    positive_words = posfile.read().lower()
positiveWordList = positive_words.split('\n')

with open(negative_dictionary, 'r') as negfile:
    negative_word = negfile.read().lower()
negativeWordList = negative_word.split('\n')


def execute():
    input_sheet = read(input_file).active
    output_wb = read(output_raw_file)
    output_sheet = output_wb.active
    for rowIndex in range(2, input_sheet.max_row + 1):
        try:
            # for rowIndex in range(2, 3):
            url_id = str(input_sheet.cell(row=rowIndex, column=1).value)
            url = input_sheet.cell(row=rowIndex, column=2).value
            if url is not None:
                article_dictionary = scrap(get_html_doc(url))
                generate_file(article_dictionary, url_id)
                filtered_words = tokenizer(article_dictionary['article_text'], read(stopwords_generic))
                positive_count = positive_score(filtered_words)
                negative_count = negative_score(filtered_words)
                sentence_length = average_sentence_length(article_dictionary['article_text'])
                complex_percentage = percentage_complex_word(article_dictionary['article_text'])
                output_sheet.cell(row=rowIndex, column=3).value = positive_count
                output_sheet.cell(row=rowIndex, column=4).value = negative_count
                output_sheet.cell(row=rowIndex, column=5).value = polarity_score(positive_count, negative_count)
                output_sheet.cell(row=rowIndex, column=6).value = subjectivity_score(len(filtered_words),
                                                                                     positive_count,
                                                                                     negative_count)
                output_sheet.cell(row=rowIndex, column=7).value = sentence_length
                output_sheet.cell(row=rowIndex, column=8).value = complex_percentage
                output_sheet.cell(row=rowIndex, column=9).value = fog_index(sentence_length, complex_percentage)
                output_sheet.cell(row=rowIndex, column=10).value = avg_words_per_sentence(
                    article_dictionary['article_text'])
                output_sheet.cell(row=rowIndex, column=11).value = syllable_count(article_dictionary['article_text'])
                output_sheet.cell(row=rowIndex, column=12).value = complex_words(article_dictionary['article_text'])
                output_sheet.cell(row=rowIndex, column=13).value = word_count(article_dictionary['article_text'])
                output_sheet.cell(row=rowIndex, column=14).value = personal_pronouns(article_dictionary['article_text'])
                output_sheet.cell(row=rowIndex, column=15).value = avg_word_length(article_dictionary['article_text'])
                output_wb.save(output_file)
        except:
            pass


def get_html_doc(url):
    response = requests.get(url, headers=headers)
    return response.text


def scrap(html_document):
    soup = BeautifulSoup(html_document, 'html.parser')
    article_title = soup.find('h1', attrs={'entry-title'})
    article_text = soup.find('div', attrs={'class': 'td-post-content'})
    return {'article_title': article_title.text, 'article_text': article_text.text}


def generate_file(article_dictionary, file_name):
    with open('output/' + file_name + '.txt', 'w') as file:
        file.write(article_dictionary['article_title'])
        file.write('\n')
        file.write(article_dictionary['article_text'])


# Calculating positive score
def positive_score(filtered):
    numPosWords = 0
    for word in filtered:
        if word.lower() in positiveWordList:
            numPosWords += 1

    sumPos = numPosWords
    return sumPos


# Calculating Negative score
def negative_score(filtered):
    num_neg_words = 0
    for word in filtered:
        if word.lower() in negativeWordList:
            num_neg_words += 1
    return num_neg_words


# Calculating polarity score
def polarity_score(positiveScore, negativeScore):
    pol_score = (positiveScore - negativeScore) / ((positiveScore + negativeScore) + 0.000001)
    return pol_score


# Calculating subjectivity score
def subjectivity_score(filtered_length, positiveScore, negativeScore):
    sub_score = (positiveScore + negativeScore) / ((filtered_length) + 0.000001)
    return sub_score


def average_sentence_length(text):
    sentence_list = sent_tokenize(text)
    tokens = sent_tokenize(text)
    totalWordCount = len(tokens)
    totalSentences = len(sentence_list)
    average_sent = 0
    if totalSentences != 0:
        average_sent = totalWordCount / totalSentences

    average_sent_length = average_sent

    return round(average_sent_length)


# Calculating percentage of complex word
# It is calculated using Percentage of Complex words = the number of complex words / the number of words

def percentage_complex_word(text):
    tokens = sent_tokenize(text)
    complexWord = 0
    complex_word_percentage = 0
    for word in tokens:
        vowels = 0
        if word.endswith(('es', 'ed')):
            pass
        else:
            for w in word:
                if w == 'a' or w == 'e' or w == 'i' or w == 'o' or w == 'u':
                    vowels += 1
            if (vowels > 2):
                complexWord += 1
    if len(tokens) != 0:
        complex_word_percentage = complexWord / len(tokens)

    return complex_word_percentage


def fog_index(averageSentenceLength, percentageComplexWord):
    return 0.4 * (averageSentenceLength + percentageComplexWord)


def avg_words_per_sentence(article_text):
    sentence_separator = '!?.'
    for p in sentence_separator:
        article_text = article_text.replace(p, '$$$')
    sentence = article_text.split('$$$')
    while "" in sentence:
        sentence.remove("")
    punc = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    for p in punc:
        article_text = article_text.replace(p, '').strip()
    words = article_text.split(' ')
    while "" in words:
        words.remove("")
    return len(words) / len(sentence)


def syllable_count(article_text):
    words = sent_tokenize(article_text)
    for word in words:
        word = word.lower()
        count = 0
        vowels = "aeiouy"
        if word[0] in vowels:
            count += 1
        for index in range(1, len(word)):
            if word[index] in vowels and word[index - 1] not in vowels:
                count += 1
        if word.endswith("e"):
            count -= 1
        if count == 0:
            count += 1
    return count


def word_count(article_text):
    tokens = nltk.word_tokenize(article_text)
    stop_words = set(stopwords.words("english"))
    filtered = []
    for w in tokens:
        if w not in stop_words:
            filtered.append(w)
    punc = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    count = 0
    for ele in filtered:
        if ele not in punc:
            count = count + 1
    return count


def complex_words(article_text):
    words = sent_tokenize(article_text)
    complex_word_count = 0
    for word in words:
        if syllable_count(word) > 2:
            complex_word_count += 1
    return complex_word_count


def personal_pronouns(article_text):
    prsnl_pronouns = ['I', 'we', 'my', 'ours', 'us']
    count = 0
    for word in article_text:
        if word in prsnl_pronouns:
            count = count + 1
    return count


def avg_word_length(article_text):
    punc = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    for p in punc:
        article_text = article_text.replace(p, '').strip()
    words = article_text.split(' ')
    return len([i for ele in words for i in ele]) / len(words)


def tokenizer(article_text, stop_word):
    stop_words = set(stop_word)
    filtered = []
    for w in article_text.split(' '):
        if w not in stop_words:
            filtered.append(w.replace('\n', ''))
    return filtered


def read(path):
    file = path.split('.')
    file_type = file[len(file) - 1]
    if 'txt' in file_type.strip():
        word_list = []
        for line in open(path):
            if line.strip() not in '':
                word_list.append(line.replace('\n', ''))
        return set(word_list)
    elif 'xlsx' in file_type.strip():
        return openpyxl.load_workbook(path)


if __name__ == "__main__":
    execute()
